#!/usr/bin/env python3
"""
Generalized OCDS-bulk importer — maps an official OCDS export (CSV or XLSX) into a COMPACT
browsable archive source. The "official bulk -> our pipeline" route for multi-year history
that live GePNIC listings can't reach (deep history is captcha-gated; we don't bypass).

Each state's export uses slightly different column names + date formats, so mapping is by a
set of OCDS field fallbacks + a dual-format date parser. Output is the same compact serving
layer for any source: per-year index shards + latest.json + status.json (engine:import, with
provenance). NO per-tender record files. DECLARED data, NOT observed -> a SEPARATE source,
kept apart from the live-observed stats/search (see compile.py + the app).

Used for: Assam (CSV, GODL) and Himachal (XLSX). Add a state = one CLI invocation.
Usage:
  python3 _import/ocds_bulk.py --source himachal_archive --label "Himachal · archive (2017-20)" \
      --file /tmp/hp_ocds.xlsx --sheet tenders --serve ./docs \
      --prov-source "..." --prov-license "..." --prov-portal "hptenders.gov.in"
"""
import argparse
import json
import os
import re
from datetime import datetime, timezone

SCHEME = "woodpecker"
_MON = {m: i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}


def _date(s):
    s = (s or "").strip()
    if not s:
        return ""
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)              # ISO (HP)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}T00:00:00Z"
    m = re.match(r"(\d{2})-([A-Za-z]{3})-(\d{4})", s)        # DD-Mon-YYYY (Assam)
    if m:
        mon = _MON.get(m.group(2).title())
        if mon:
            return f"{m.group(3)}-{mon:02d}-{int(m.group(1)):02d}T00:00:00Z"
    return ""


def _amount(s):
    d = re.sub(r"[^0-9.]", "", str(s or ""))
    if not d:
        return None
    try:
        return int(round(float(d)))
    except ValueError:
        return None


def load_rows(path, sheet=None):
    if path.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = [str(c).strip() if c is not None else "" for c in next(it)]
        for row in it:
            yield {hdr[i]: ("" if v is None else str(v)) for i, v in enumerate(row) if i < len(hdr)}
    else:
        import csv
        with open(path, encoding="utf-8", errors="replace") as f:
            for r in csv.DictReader(f):
                yield r


def g(r, *keys):
    for k in keys:
        v = r.get(k)
        if v not in (None, ""):
            return v.strip() if isinstance(v, str) else str(v)
    return ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True)
    ap.add_argument("--label", required=True)
    ap.add_argument("--file", required=True)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--serve", default="./docs")
    ap.add_argument("--prov-source", default="official OCDS bulk export")
    ap.add_argument("--prov-license", default="")
    ap.add_argument("--prov-portal", default="")
    args = ap.parse_args()

    SOURCE = args.source
    out = os.path.join(args.serve, SOURCE)
    os.makedirs(os.path.join(out, "index"), exist_ok=True)
    imported_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    entries, shards, seen = [], {}, set()
    for r in load_rows(args.file, args.sheet):
        tid = g(r, "tender/id", "id").strip()
        if not tid or tid in seen:
            continue
        seen.add(tid)
        pub = _date(g(r, "tender/datePublished", "date"))
        val = _amount(g(r, "tender/value/amount"))
        e = {
            "ocid": f"{SCHEME}-{SOURCE}-{tid}",
            "source": SOURCE,
            "title": (g(r, "tender/title") or tid),
            "buyer": g(r, "buyer/name", "tender/procuringEntity/name", "parties/0/contactPoint/name"),
            "category": g(r, "tender/mainProcurementCategory").lower(),
            "value": ({"amount": val, "currency": "INR"} if val else {}),
            "status": g(r, "tender/status").lower() or "active",
            "publishedDate": pub,
            "closingDate": _date(g(r, "tender/tenderPeriod/endDate")),
            "latestTag": "tender",
            "lastReleaseDate": pub,
            "releaseCount": 1, "corrigendumCount": 0,
            "hasDeadLink": False, "hasOcr": False,
            "hasUndeclaredChange": False, "pulled": False,
            "observationCount": 0, "imported": True,
            "fiscalYear": g(r, "fiscal_year"),
            "reference": g(r, "tender/externalReference"),
        }
        entries.append(e)
        shards.setdefault(pub[:4] or e["fiscalYear"] or "unknown", []).append(e)

    def wj(path, obj):
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(obj, fh, ensure_ascii=False)

    for yr, items in shards.items():
        items.sort(key=lambda e: e["publishedDate"], reverse=True)
        wj(os.path.join(out, "index", f"{yr}.json"),
           {"source": SOURCE, "month": yr, "count": len(items), "tenders": items})
    latest = sorted(entries, key=lambda e: e["publishedDate"], reverse=True)[:200]
    wj(os.path.join(out, "index", "latest.json"),
       {"source": SOURCE, "month": "latest", "count": len(latest),
        "months": sorted(shards.keys(), reverse=True), "tenders": latest})
    wj(os.path.join(out, "status.json"),
       {"source": SOURCE, "label": args.label, "engine": "import", "ok": True,
        "mode": "import", "lastRun": imported_at, "importedAt": imported_at,
        "tenders": len(entries), "tendersCaptured": len(entries),
        "provenance": {"imported": True, "source": args.prov_source,
                       "sourcePortal": args.prov_portal, "license": args.prov_license,
                       "note": "Declared official export — NOT observed by Woodpecker (no capture-floor snapshots)."},
        "lastError": None})
    print(f"[{SOURCE}] imported {len(entries)} tenders across {len(shards)} year shard(s) -> {out}  (years: {', '.join(sorted(shards))})")


if __name__ == "__main__":
    main()
